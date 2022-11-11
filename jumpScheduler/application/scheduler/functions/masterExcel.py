# functions.py

import sys
import os
import xlsxwriter
import shutil
from copy import deepcopy
from openpyxl import load_workbook


# Get import path information
current = os.path.dirname(os.path.realpath(__file__))
parent = os.path.dirname(current)
sys.path.append(parent)
 
# Import from parent directory
import settings as settings
from classes import classes as classes
import functions.opperations as operations


def readInClassesFromExcell(filename):

    allClassesDic = {}

    #load excel file
    workbook = load_workbook(filename)
    
    # Get main sheet 
    sheetToGet = 'Sections.export'

    for sheet in range(len(workbook.sheetnames)):
        if workbook.sheetnames[sheet] == sheetToGet:
            break

    workbook.active = sheet
    
    teamSheet = workbook.active

    # Update the cores for each student in the master excell sheet 
    for i in range(2, 400):

        schoolID = str(teamSheet[f'A{i}'].value)

        if schoolID:

            # Create row from spreadsheet
            row = [schoolID, 
            str(teamSheet[f'B{i}'].value), 
            str(teamSheet[f'C{i}'].value), 
            str(teamSheet[f'D{i}'].value), 
            str(teamSheet[f'E{i}'].value),
            str(teamSheet[f'F{i}'].value),
            str(teamSheet[f'G{i}'].value),
            str(teamSheet[f'H{i}'].value),
            str(teamSheet[f'I{i}'].value)]
            
            # Add classes to dictionary with section numbers as their Keys
            allClassesDic[row[1]] = classes.classSection(row)

        else:

            break

    return allClassesDic

def readInStudentsFromExcell(filename):

    #load excel file
    workbook = load_workbook(filename)
    
    # Get main sheet 
    workbook.active = workbook.worksheets[0]
    
    sheetOne = workbook.active


    allStuOneDic = {}

    allStuTwoDic = {}

    allStuFinalDIc = {}



     # Update the cores for each student in the master excell sheet 
    for i in range(2, 130):

        stuID = str(sheetOne[f'C{i}'].value)

        if stuID and stuID[0] != '#':

            # Create row from spreadsheet
            row = [
            str(sheetOne[f'A{i}'].value), 
            str(sheetOne[f'B{i}'].value), 
            stuID, 
            str(sheetOne[f'D{i}'].value), 
            str(sheetOne[f'E{i}'].value),
            str(sheetOne[f'F{i}'].value),
            str(sheetOne[f'G{i}'].value),
            str(sheetOne[f'H{i}'].value),
            str(sheetOne[f'I{i}'].value),
            str(sheetOne[f'J{i}'].value),
            str(sheetOne[f'K{i}'].value),
            str(sheetOne[f'L{i}'].value),
            str(sheetOne[f'M{i}'].value),
            str(sheetOne[f'N{i}'].value),
            str(sheetOne[f'O{i}'].value),
            str(sheetOne[f'P{i}'].value),
            str(sheetOne[f'Q{i}'].value),
            str(sheetOne[f'R{i}'].value),
            str(sheetOne[f'S{i}'].value)
            ]

            row[:] = [x if x != '\xa0' else '' for x in row]
            row[:] = [x if x != 'None' else '' for x in row]
            
            # Add students to dictionary with ID as their Keys
            allStuOneDic[row[2]] = row


    # Get studxents grade sheets info
    sheetsToGet = ['7th All Info', '8th All Info']

    for sheetToGet in sheetsToGet:
        
        for sheet in range(len(workbook.sheetnames)):
         if workbook.sheetnames[sheet].casefold() == sheetToGet.casefold():
            break

        workbook.active = sheet
    
        sheetTwo = workbook.active

        # Check for extra column in sheet 
        if "Academic" not in str(sheetTwo['T1'].value): raise Exception("ERROR: Please delete column T in 7th or 8th info")

        # Parse the 8th and 7th info sheets 
        for i in range(2, 60):

            row = []

            for cell in sheetTwo[i]:
                row.append(str(cell.value))

            row[:] = [x if x != '\xa0' else '' for x in row]
            row[:] = [x if x != 'None' else '' for x in row]

            if row[0]:

                allStuTwoDic[row[2]] = row

    
    #emptyStuArr = ['Magana', 'Asher', '39054', '6', 'M', 'Pryor', 'Cascades', '8/17/09', '', '', '', '', '', '', '', '', '', '', '', '2', '2', '7th', 'Grade', 'Grade', 'Level 5', 'Level 3', '', 'ASC', '', '', '', '', 'Recommend', '', 'Vincent Jarvis, Dominic Flucas ', 'Devin Box', '', '']
    emptyStuArr = ['', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '']


    for stuOne in allStuOneDic:

        if stuOne in allStuTwoDic:

            allStuFinalDIc[stuOne] = classes.student(allStuTwoDic[stuOne])
            
        elif stuOne:

            newStu = deepcopy(emptyStuArr)

            newStu[0] = allStuOneDic[stuOne][0]
            newStu[1] = allStuOneDic[stuOne][1]
            newStu[2] = allStuOneDic[stuOne][2]
            newStu[3] = str(int(allStuOneDic[stuOne][4]) - 1)
            newStu[4] = allStuOneDic[stuOne][3]

            if allStuOneDic[stuOne][5] != '': newStu[9] = 'Yes' # IEP
            if allStuOneDic[stuOne][6] != '': newStu[17] = 'Yes' # TAG

            if allStuOneDic[stuOne][7] != '':
                if 'IEP' in allStuOneDic[stuOne][7]:
                    if 'M' in allStuOneDic[stuOne][7]:
                        newStu[12] = 'Current'
                    if 'R' in allStuOneDic[stuOne][7]:
                        newStu[10] = 'Current'
                    if 'W' in allStuOneDic[stuOne][7]:
                        newStu[11] = 'Current'
                if 'Speech' in allStuOneDic[stuOne][7]:
                    newStu[13] = 'Current'

            newStu[19] = allStuOneDic[stuOne][8] # Acidem rate
            newStu[20] = allStuOneDic[stuOne][9] # Behave rate

            newStu[21] = allStuOneDic[stuOne][10] # Math plcment
            
            newStu[22] = allStuOneDic[stuOne][15] # Reading 
            newStu[23] = allStuOneDic[stuOne][16] # Writing 

            newStu[24] = allStuOneDic[stuOne][11] # ireadela 
            newStu[25] = allStuOneDic[stuOne][12] # ireadymath

            allStuFinalDIc[stuOne] = classes.student(newStu)

        if stuOne:

            # spanish placement
            if allStuOneDic[stuOne][-1] == "SP Adv": 
                allStuFinalDIc[stuOne].spanishPlacement = "M_SPN_ADV"
                allStuFinalDIc[stuOne].requestedClasses.append('M_SPN_ADV')
            else: 
                allStuFinalDIc[stuOne].spanishPlacement = "M_SPN"
                allStuFinalDIc[stuOne].requestedClasses.append('M_SPN')
    

                
    return allStuFinalDIc

def getClassCode(unparsedClass):
    """ function to get the class code 
    """

    if 'SS' in unparsedClass:

        return 'SS'

    elif 'SCI' in unparsedClass:

        return 'SCI'

    elif 'MATH7' in unparsedClass:

        return 'IM7'

    elif 'MATH8' in unparsedClass:

        return 'IM8'

    elif 'SPN_ADV' in unparsedClass:

        return 'SP Adv'
    
    elif 'SP' in unparsedClass:

        return 'SP'

    elif 'LA' in unparsedClass:

        return 'LA'

def writeToMasterExcell(teamStudentsDic):
    """ function to write scheduled students to master excell sheet 
    """

    print("Begining Excel Write")

    # Get the master excell file 
    file_name = os.listdir('application/excell/import/')

    # duplicate the file 
    shutil.copy2(f'application/excell/import/{file_name[0]}', 'application/excell/export/')

    # load excel file
    # workbookRead = load_workbook(filename=f'application/excell/import/{file_name[0]}', data_only=True)
    workbook = load_workbook(f'application/excell/export/{file_name[0]}')
    
    # Get main sheet 
    # worksheetRead = workbookRead.worksheets[0]
    workbook.active = workbook.worksheets[0]
    
    teamSheet = workbook.active


    # Initial call to print 0% progress
    print('EDITING EXCELL SHEET')
    #operations.printProgressBar(0, 130, prefix = 'Progress:', suffix = 'Complete', length = 50)

    # Update the cores for each student in the master excell sheet 
    for i in range(2, 130):

        # Update Progress Bar
        #operations.printProgressBar(i - 1, 130, prefix = 'Progress:', suffix = 'Complete', length = 50)

        stuID = str(teamSheet[f'C{i}'].value)

        if stuID and stuID[0] != '#' and stuID in teamStudentsDic:

            # modify the students classes cells

            # Core 1
            if teamStudentsDic[stuID].studentSched['_1B']:
                cOne = getClassCode(teamStudentsDic[stuID].studentSched['_1B'].courseNumber)
            else: cOne = ''
            teamSheet[f'T{i}'] = cOne

            # Core 2
            if teamStudentsDic[stuID].studentSched['_2B']:
                cTwo = getClassCode(teamStudentsDic[stuID].studentSched['_2B'].courseNumber)
            else: cTwo = ''
            teamSheet[f'U{i}'] = cTwo

            # Core 3
            if teamStudentsDic[stuID].studentSched['_1G']:
                cThree = getClassCode(teamStudentsDic[stuID].studentSched['_1G'].courseNumber)
            else: cThree = ''
            teamSheet[f'V{i}'] = cThree

            # Core 4
            if teamStudentsDic[stuID].studentSched['_2G']:
                cFour = getClassCode(teamStudentsDic[stuID].studentSched['_2G'].courseNumber)
            else: cFour = ''
            teamSheet[f'W{i}'] = cFour

            # Core 5
            if teamStudentsDic[stuID].studentSched['_5G']:
                cFive = getClassCode(teamStudentsDic[stuID].studentSched['_5G'].courseNumber)
            else: cFive = ''
            teamSheet[f'X{i}'] = cFive

        
        else:

            if stuID != "None":
                print(f'ERROR: Student {stuID} in master excell sheet not found in scheduled students dictionary')        
        
    
        #save the file
        workbook.save(f'application/excell/export/{file_name[0]}')

        # workbookRead.close()
        workbook.close()

        print("EDITING FINISHED")

